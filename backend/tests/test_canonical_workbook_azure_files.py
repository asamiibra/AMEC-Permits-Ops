from pathlib import Path

import pytest
from openpyxl import Workbook as OpenpyxlWorkbook
from openpyxl import load_workbook as real_load_workbook

from backend.app.adapters.excel import adapter as adapter_module
from backend.app.fixtures.canonical import (
    CANONICAL_PROJECTION_SHEET,
    CANONICAL_PROJECT_IDS,
    canonical_workbook_path,
)
from backend.app.services import canonical_workbook as module


def _prepare(monkeypatch, tmp_path):
    root = tmp_path / "workspace"
    monkeypatch.setenv("SYNTHETIC_TEST_ROOT", str(root))
    monkeypatch.setenv("MOCK_SYSTEMS_ROOT", str(root / "mock-systems"))
    path = canonical_workbook_path()
    module.ensure_canonical_workbook(path)
    return path


def _human_values(path):
    workbook = real_load_workbook(path, data_only=False)
    try:
        return {sheet: tuple(tuple(cell.value for cell in row) for row in workbook[sheet].iter_rows()) for sheet in module.HUMAN_HEADERS}
    finally:
        workbook.close()


def _write(path, project=CANONICAL_PROJECT_IDS[0], status="SEEDED"):
    return module.write_system_projection(path, project, {
        "Canonical Plot Number": "001234",
        "Canonical PIN": "PIN-000123",
        "Rendering Version": "R1.0",
        "Municipality Request": "GHCE-APP-0142",
        "Projection Status": status,
    })


class TrackingWorkbook:
    def __init__(self, workbook, *, fail_sheetnames=False, fail_getitem=False, fail_save=False):
        self._workbook = workbook
        self.closed = False
        self.fail_sheetnames = fail_sheetnames
        self.fail_getitem = fail_getitem
        self.fail_save = fail_save

    @property
    def sheetnames(self):
        if self.fail_sheetnames:
            raise RuntimeError("synthetic workbook inspection failure")
        return self._workbook.sheetnames

    def __getitem__(self, key):
        if self.fail_getitem:
            raise RuntimeError("synthetic workbook sheet failure")
        return self._workbook[key]

    def save(self, *args, **kwargs):
        if self.fail_save:
            raise OSError("synthetic workbook save failure")
        return self._workbook.save(*args, **kwargs)

    def close(self):
        self.closed = True
        return self._workbook.close()

    def __getattr__(self, name):
        return getattr(self._workbook, name)


def _tracked_loader(trackers, *, fail_sheetnames=False, fail_getitem=False):
    def loader(path, *args, **kwargs):
        tracked = TrackingWorkbook(real_load_workbook(path, *args, **kwargs), fail_sheetnames=fail_sheetnames, fail_getitem=fail_getitem)
        trackers.append(tracked)
        return tracked
    return loader


def test_atomic_write_preserves_human_sheets_and_reports_hashes(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    before = _human_values(path)
    result = _write(path)
    assert result["write_mode"] == "ATOMIC_REPLACE"
    assert result["candidate_sha256"] == result["destination_sha256"]
    assert _human_values(path) == before


def test_four_canonical_projects_write_sequentially(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    for index, project in enumerate(CANONICAL_PROJECT_IDS):
        result = module.write_system_projection(path, project, {
            "Canonical Plot Number": f"00{index + 1234}",
            "Canonical PIN": f"PIN-{index:06d}",
            "Rendering Version": "R1.0",
            "Municipality Request": f"GHCE-APP-{index + 142:04d}",
            "Projection Status": "SEEDED",
        })
        assert result["write_mode"] == "ATOMIC_REPLACE"
    workbook = real_load_workbook(path, data_only=True)
    try:
        sheet = workbook[CANONICAL_PROJECTION_SHEET]
        assert [sheet.cell(row, 1).value for row in range(2, 6)] == CANONICAL_PROJECT_IDS
        assert [sheet.cell(row, 6).value for row in range(2, 6)] == ["SEEDED"] * 4
    finally:
        workbook.close()


def test_replace_is_same_directory_and_attempted_once(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    calls = []
    original_replace = module.os.replace

    def replace(source, destination):
        calls.append((Path(source), Path(destination)))
        return original_replace(source, destination)

    monkeypatch.setattr(module.os, "replace", replace)
    _write(path)
    assert len(calls) == 1
    assert calls[0][0].parent == path.parent
    assert calls[0][1] == path


def test_candidate_fsync_is_once_after_close_before_validation_and_replace(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    events = []
    trackers = []
    original_fsync_candidate = module._fsync_candidate
    original_validate = module._validate_workbook_candidate
    original_replace = module.os.replace

    def tracked_fsync_candidate(candidate):
        events.append(("fsync_candidate", Path(candidate), trackers[0].closed))
        return original_fsync_candidate(candidate)

    def tracked_validate(candidate, *args):
        events.append(("validate", Path(candidate)))
        return original_validate(candidate, *args)

    def tracked_replace(candidate, destination):
        events.append(("replace", Path(candidate), Path(destination)))
        return original_replace(candidate, destination)

    monkeypatch.setattr(module, "load_workbook", _tracked_loader(trackers))
    monkeypatch.setattr(module, "_fsync_candidate", tracked_fsync_candidate)
    monkeypatch.setattr(module, "_validate_workbook_candidate", tracked_validate)
    monkeypatch.setattr(module.os, "replace", tracked_replace)
    _write(path)

    fsync_events = [event for event in events if event[0] == "fsync_candidate"]
    validate_events = [event for event in events if event[0] == "validate"]
    replace_events = [event for event in events if event[0] == "replace"]
    assert len(fsync_events) == 1
    assert len(validate_events) == 2
    assert len(replace_events) == 1
    assert fsync_events[0][1] != path
    assert fsync_events[0][1].parent == path.parent
    assert fsync_events[0][2] is True
    assert events.index(fsync_events[0]) < events.index(validate_events[0]) < events.index(replace_events[0])


def test_candidate_fsync_failure_is_fail_closed_and_preserves_original(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    before = path.read_bytes()
    replace_calls = []
    fsync_error = OSError("synthetic candidate fsync failure")

    monkeypatch.setattr(module.os, "fsync", lambda _descriptor: (_ for _ in ()).throw(fsync_error))
    monkeypatch.setattr(module.os, "replace", lambda *args: replace_calls.append(args))
    with pytest.raises(OSError, match="synthetic candidate fsync failure") as raised:
        _write(path)

    assert raised.value is fsync_error
    assert path.read_bytes() == before
    assert replace_calls == []
    assert not list(path.parent.glob("permitops-excel-*.xlsx"))


def test_permission_error_is_fail_closed_and_preserves_original(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    before = path.read_bytes()
    monkeypatch.setattr(module.os, "replace", lambda *_args: (_ for _ in ()).throw(PermissionError(13, "Permission denied")))
    with pytest.raises(module.WorkbookWriteError, match="ATOMIC_WORKBOOK_REPLACE=FAIL") as raised:
        _write(path)
    assert isinstance(raised.value.__cause__, PermissionError)
    assert path.read_bytes() == before
    assert not list(path.parent.glob("permitops-excel-*.xlsx"))


def test_invalid_candidate_leaves_destination_unchanged_and_does_not_replace(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    before = path.read_bytes()
    replace_calls = []
    monkeypatch.setattr(module.os, "replace", lambda *args: replace_calls.append(args))
    monkeypatch.setattr(module, "_validate_workbook_candidate", lambda *args: (_ for _ in ()).throw(module.WorkbookWriteError("TEMP_WORKBOOK_VALIDATION=FAIL")))
    with pytest.raises(module.WorkbookWriteError, match="TEMP_WORKBOOK_VALIDATION=FAIL"):
        _write(path)
    assert path.read_bytes() == before
    assert replace_calls == []
    assert not list(path.parent.glob("permitops-excel-*.xlsx"))


def test_cleanup_failure_never_masks_primary_failure(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    monkeypatch.setattr(module, "_validate_workbook_candidate", lambda *args: (_ for _ in ()).throw(module.WorkbookWriteError("TEMP_WORKBOOK_VALIDATION=FAIL")))
    original_unlink = Path.unlink

    def unlink(self, *args, **kwargs):
        if self.name.startswith("permitops-excel-"):
            raise OSError("synthetic cleanup failure")
        return original_unlink(self, *args, **kwargs)

    monkeypatch.setattr(Path, "unlink", unlink)
    with pytest.raises(module.WorkbookWriteError, match="TEMP_WORKBOOK_VALIDATION=FAIL"):
        _write(path)


def test_temporary_candidate_is_removed_after_success(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    _write(path)
    assert not list(path.parent.glob("permitops-excel-*.xlsx"))


def test_temporary_candidate_is_removed_after_replace_failure(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    monkeypatch.setattr(module.os, "replace", lambda *_args: (_ for _ in ()).throw(OSError("synthetic replace failure")))
    with pytest.raises(OSError, match="synthetic replace failure"):
        _write(path)
    assert not list(path.parent.glob("permitops-excel-*.xlsx"))


def test_lock_behavior_is_unchanged(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    lock = path.with_suffix(path.suffix + ".lock")
    lock.write_text("synthetic lock", encoding="utf-8")
    try:
        with pytest.raises(module.WorkbookLockedError, match="MANUAL_COPY_REQUIRED"):
            _write(path)
    finally:
        lock.unlink()


def test_contract_closes_read_only_workbook_on_success(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    trackers = []
    monkeypatch.setattr(module, "load_workbook", _tracked_loader(trackers))
    contract = module.canonical_workbook_contract(path)
    assert CANONICAL_PROJECTION_SHEET in contract["sheets"]
    assert trackers and all(item.closed for item in trackers)


def test_contract_closes_read_only_workbook_on_exception(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    trackers = []
    monkeypatch.setattr(module, "load_workbook", _tracked_loader(trackers, fail_sheetnames=True))
    with pytest.raises(RuntimeError, match="inspection failure"):
        module.canonical_workbook_contract(path)
    assert trackers and all(item.closed for item in trackers)


def test_candidate_validation_closes_workbook_on_success(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    _write(path)
    human_snapshot = _human_values(path)
    trackers = []
    monkeypatch.setattr(module, "load_workbook", _tracked_loader(trackers))
    module._validate_workbook_candidate(
        path,
        CANONICAL_PROJECT_IDS[0],
        {"Projection Status": "SEEDED"},
        human_snapshot,
    )
    assert trackers and all(item.closed for item in trackers)


def test_candidate_validation_closes_workbook_on_exception(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    trackers = []
    monkeypatch.setattr(module, "load_workbook", _tracked_loader(trackers, fail_getitem=True))
    with pytest.raises(module.WorkbookWriteError, match="TEMP_WORKBOOK_VALIDATION=FAIL"):
        module._validate_workbook_candidate(path, CANONICAL_PROJECT_IDS[0], {}, _human_values(path))
    assert trackers and all(item.closed for item in trackers)


def test_adapter_read_rows_closes_on_success(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    trackers = []
    monkeypatch.setattr(adapter_module, "load_workbook", _tracked_loader(trackers))
    rows = adapter_module.MockExcelAdapter(str(path)).read_rows()
    assert rows and trackers and all(item.closed for item in trackers)


def test_adapter_read_rows_closes_on_exception(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    trackers = []
    monkeypatch.setattr(adapter_module, "load_workbook", _tracked_loader(trackers, fail_getitem=True))
    with pytest.raises(RuntimeError, match="sheet failure"):
        adapter_module.MockExcelAdapter(str(path)).read_rows()
    assert trackers and all(item.closed for item in trackers)


def test_adapter_resolve_row_identity_closes_on_matching_return(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    trackers = []
    monkeypatch.setattr(adapter_module, "load_workbook", _tracked_loader(trackers))
    result = adapter_module.MockExcelAdapter(str(path)).resolve_row_identity(CANONICAL_PROJECT_IDS[0])
    assert result["row_number"] == 2
    assert trackers and all(item.closed for item in trackers)


def test_adapter_resolve_row_identity_closes_on_not_found(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    trackers = []
    monkeypatch.setattr(adapter_module, "load_workbook", _tracked_loader(trackers))
    assert adapter_module.MockExcelAdapter(str(path)).resolve_row_identity("NOT-FOUND") is None
    assert trackers and all(item.closed for item in trackers)


def test_adapter_resolve_row_identity_closes_on_exception(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    trackers = []
    monkeypatch.setattr(adapter_module, "load_workbook", _tracked_loader(trackers, fail_getitem=True))
    with pytest.raises(RuntimeError, match="sheet failure"):
        adapter_module.MockExcelAdapter(str(path)).resolve_row_identity(CANONICAL_PROJECT_IDS[0])
    assert trackers and all(item.closed for item in trackers)


def test_ensure_workbook_closes_when_save_succeeds(monkeypatch, tmp_path):
    tracked = TrackingWorkbook(OpenpyxlWorkbook())
    monkeypatch.setattr(module, "Workbook", lambda: tracked)
    module.ensure_canonical_workbook(tmp_path / "created.xlsx")
    assert tracked.closed


def test_ensure_workbook_closes_when_save_raises(monkeypatch, tmp_path):
    tracked = TrackingWorkbook(OpenpyxlWorkbook(), fail_save=True)
    monkeypatch.setattr(module, "Workbook", lambda: tracked)
    with pytest.raises(OSError, match="save failure"):
        module.ensure_canonical_workbook(tmp_path / "created.xlsx")
    assert tracked.closed


def test_writable_workbook_closes_when_save_raises(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    trackers = []

    def loader(path_value, *args, **kwargs):
        tracked = TrackingWorkbook(real_load_workbook(path_value, *args, **kwargs), fail_save=True)
        trackers.append(tracked)
        return tracked

    monkeypatch.setattr(module, "load_workbook", loader)
    with pytest.raises(OSError, match="save failure"):
        _write(path)
    assert trackers and trackers[0].closed


def test_post_replace_validation_workbooks_close(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    trackers = []
    monkeypatch.setattr(module, "load_workbook", _tracked_loader(trackers))
    result = _write(path)
    assert result["destination_sha256"] == result["candidate_sha256"]
    assert len(trackers) >= 3
    assert all(item.closed for item in trackers)


def test_permission_error_has_no_environment_fallback(monkeypatch, tmp_path):
    path = _prepare(monkeypatch, tmp_path)
    monkeypatch.setattr(module.os, "replace", lambda *_args: (_ for _ in ()).throw(PermissionError("replace denied")))
    with pytest.raises(module.WorkbookWriteError, match="ATOMIC_WORKBOOK_REPLACE=FAIL"):
        _write(path)


def test_removed_recovery_path_and_exception_are_not_reachable():
    source = Path(module.__file__).read_text(encoding="utf-8")
    assert "_bounded_azure_files_overwrite" not in source
    assert "_restore_original_workbook" not in source
    assert "_write_full_file" not in source
    assert "WorkbookWriteRecoveryError" not in source
    assert "SYNTHETIC_AZURE_FILES_BOUNDED_OVERWRITE" not in source
