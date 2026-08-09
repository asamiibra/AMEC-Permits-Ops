import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook

from backend.app.adapters.excel.adapter import MockExcelAdapter, WorkbookLockedError
from backend.app.fixtures.canonical import CANONICAL_FIXTURE_MANIFEST, CANONICAL_WORKBOOK, CANONICAL_WORKBOOK_SHEETS, fixture_metadata
from backend.app.services.canonical_workbook import canonical_workbook_contract


def canonical_project(client):
    return next(x for x in client.get("/api/projects").json() if x["project_number"] == "GHCE-2026-0142")


def test_canonical_fixture_manifest_is_authoritative_and_deterministic(client):
    body = client.get("/api/reconciliation/fixture").json()
    expected_hash = hashlib.sha256(json.dumps(CANONICAL_FIXTURE_MANIFEST, sort_keys=True, separators=(",", ":")).encode()).hexdigest()
    assert body["fixture_set_id"] == "PermitOps_Synthetic_MVP_Dataset_v1"
    assert body["status"] == "ACTIVE_GOLDEN_PATH"
    assert body["manifest_sha256"] == expected_hash
    assert len(body["manifest"]["portal_attachment_categories"]) == 17
    assert len(body["manifest"]["undertaking_authorization_forms"]) == 6
    assert client.get("/api/evaluation/analysis").json()["profile"]["fixture_set"] == body["fixture_set_id"]


def test_recording_derived_excel_contract_and_safe_projection(client):
    contract = client.get("/api/reconciliation/excel-contract").json()
    assert all(sheet in contract["sheets"] for sheet in CANONICAL_WORKBOOK_SHEETS)
    assert contract["system_owned_projection"]["sheet"] == "PERMITOPS SYSTEM PROJECTION"
    project = canonical_project(client)
    path = Path(CANONICAL_WORKBOOK)
    before = load_workbook(path, data_only=False)
    human_before = before["GENERAL FOLLOW UP"]["F2"].value
    result = client.post(f"/api/reconciliation/excel-projection/{project['id']}", json={"canonical_plot_number":"001234","canonical_pin":"PIN-000123"})
    assert result.status_code == 200 and result.json()["human_owned_cells_changed"] is False
    after = load_workbook(path, data_only=False)
    assert after["GENERAL FOLLOW UP"]["F2"].value == human_before
    assert after["PERMITOPS SYSTEM PROJECTION"]["B2"].value == "001234"


def test_locked_workbook_is_a_controlled_exception():
    path = Path(CANONICAL_WORKBOOK)
    lock = path.with_suffix(path.suffix + ".lock")
    lock.write_text("synthetic lock", encoding="utf-8")
    try:
        try:
            MockExcelAdapter(str(path)).write_system_projection("GHCE-2026-0142", {"Canonical Plot Number":"009999"})
            assert False, "locked workbook write should fail"
        except WorkbookLockedError as error:
            assert "MANUAL_COPY_REQUIRED" in str(error)
    finally:
        lock.unlink(missing_ok=True)


def test_multi_owner_property_representation_and_evidence(client):
    project = canonical_project(client)
    body = client.get(f"/api/reconciliation/properties/{project['id']}").json()
    assert body["property"]["plot_number"] == "001234"
    assert len(body["owners"]) == 2
    assert sum(owner["normalized_share"] for owner in body["owners"]) == 1.0
    assert all(owner["party"]["identifier_value"].startswith("QID-") for owner in body["owners"])
    assert body["representations"]
    assert body["representations"][0]["representative"]["id"] not in {owner["party"]["id"] for owner in body["owners"]}
    assert body["authorizations"][0]["status"] == "VALID"
    assert body["source_evidence"]["document_version_id"]


def test_target_rendering_is_target_neutral_and_versioned(client):
    project = canonical_project(client)
    body = client.get(f"/api/reconciliation/rendering-preview?project_id={project['id']}&field_code=PROPERTY.PLOT_NUMBER").json()
    assert body["raw_observation"]["raw_value"] == body["canonical_verified_value"]["value"]
    assert body["canonical_verified_value"]["target_neutral"] is True
    assert set(body["target_renderings"]) == {"FORM", "EXCEL", "MUNICIPALITY"}
    assert all(item["rule_version"] == "1.0" for item in body["target_renderings"].values())


def test_project_bootstrap_chain_and_number_conflict(client):
    response = client.post("/api/reconciliation/project-bootstrap", json={"initiation_reference":"SYNTHETIC-BOOTSTRAP-TEST","project_name":"Synthetic Bootstrap Test"})
    assert response.status_code == 200
    body = response.json()
    assert body["project"]["project_number"] == "GHCE-2026-0245"
    assert body["synology"]["template_applied"] is True
    assert body["excel"]["projection_sheet"] == "PERMITOPS SYSTEM PROJECTION"
    assert body["application"]["external_request_number"].startswith("REQ-BOOTSTRAP-")
    assert {"PROJECT_INITIATED","PROJECT_NUMBER_RESERVED","PROJECT_CREATED","SYNOLOGY_PROJECT_ROOT_CREATED","PROJECT_TEMPLATE_APPLIED","EXCEL_PROJECT_ROW_LINKED"}.issubset(set(body["audit"]))
    duplicate = client.post("/api/reconciliation/project-bootstrap", json={"initiation_reference":"SYNTHETIC-BOOTSTRAP-DUP","project_name":"Duplicate Number","proposed_number":"GHCE-2026-0245"})
    assert duplicate.status_code == 409


def test_conflicts_have_explicit_synthetic_dispositions(client):
    project = canonical_project(client)
    conflicts = client.get(f"/api/projects/{project['id']}/conflicts").json()
    assert all(conflict["status"] in {"RESOLVED", "ACCEPTED"} for conflict in conflicts)
