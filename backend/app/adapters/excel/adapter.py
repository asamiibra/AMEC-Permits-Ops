from pathlib import Path
from typing import Protocol
from openpyxl import load_workbook
from ...services.canonical_workbook import canonical_workbook_contract, write_system_projection, WorkbookLockedError


class ExcelAdapter(Protocol):
    def read_rows(self) -> list[dict]: ...
    def get_project_row(self, project_number: str) -> dict | None: ...
    def resolve_row_identity(self, project_number: str) -> dict | None: ...
    def write_system_projection(self, project_number: str, values: dict) -> dict: ...
    def health_check(self) -> dict: ...


class MockExcelAdapter:
    def __init__(self, path: str): self.path = Path(path)

    def read_rows(self) -> list[dict]:
        wb = load_workbook(self.path, read_only=True, data_only=True)
        try:
            ws = wb["GENERAL FOLLOW UP"]
            rows = list(ws.iter_rows(values_only=True))
            headers = [str(v) for v in rows[0]]
            return [dict(zip(headers, row)) for row in rows[1:] if any(row)]
        finally:
            wb.close()

    def get_project_row(self, project_number: str) -> dict | None:
        return next((row for row in self.read_rows() if row.get("Project Number") == project_number), None)

    def resolve_row_identity(self, project_number: str) -> dict | None:
        wb = load_workbook(self.path, read_only=True, data_only=True)
        try:
            ws = wb["GENERAL FOLLOW UP"]
            headers = [cell.value for cell in ws[1]]
            project_index = headers.index("Project Number") + 1
            for row_number in range(2, ws.max_row + 1):
                if ws.cell(row_number, project_index).value == project_number:
                    return {"workbook_identity": str(self.path), "sheet_name": "GENERAL FOLLOW UP", "row_number": row_number, "row_key": project_number}
            return None
        finally:
            wb.close()

    def write_system_projection(self, project_number: str, values: dict) -> dict:
        return write_system_projection(self.path, project_number, values)

    def contract(self) -> dict:
        return canonical_workbook_contract(self.path)

    def health_check(self) -> dict:
        result = {"adapter": "EXCEL", "status": "OK" if self.path.exists() else "UNAVAILABLE", "path": str(self.path), "synthetic": True, "write_policy": "SYSTEM_PROJECTION_ONLY"}
        if self.path.exists():
            result["contract"] = self.contract()
        return result
