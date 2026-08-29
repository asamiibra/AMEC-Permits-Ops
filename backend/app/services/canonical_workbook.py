from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any
import os
import hashlib
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from ..config.settings import get_settings
from ..fixtures.canonical import CANONICAL_PROJECTION_SHEET, CANONICAL_WORKBOOK_SHEETS, canonical_workbook_path
from ..fixtures.canonical import CANONICAL_PROJECT_IDS


PROJECTION_HEADERS = ["Project Number", "Canonical Plot Number", "Canonical PIN", "Rendering Version", "Municipality Request", "Projection Status"]
HUMAN_HEADERS = {
    "GENERAL FOLLOW UP": ["Project Number", "Project Name", "Client/Owner", "Status", "Permit Type", "Human Notes"],
    "DESIGN": ["Project Number", "Drawing Revision", "Design Lead", "Human Notes"],
    "Suppervission": ["Project Number", "Supervisor", "Review Status", "Human Notes"],
    "Services Provider": ["Project Number", "Provider", "Service Type", "Human Notes"],
}
HUMAN_ROWS = [
    (CANONICAL_PROJECT_IDS[0], "Al Noor Villa", "Synthetic Owner Group", "DRAFT", "Building Permit"),
    (CANONICAL_PROJECT_IDS[1], "West Bay Residence", "Synthetic Owner Group", "RETURNED", "Building Permit"),
    (CANONICAL_PROJECT_IDS[2], "Lusail Office Annex", "Synthetic Company", "UNDER_REVIEW", "Fit-out Permit"),
    (CANONICAL_PROJECT_IDS[3], "Pearl Community Clinic", "Synthetic Company", "APPROVED", "Renovation Permit"),
]


def ensure_canonical_workbook(path: str | Path) -> Path:
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    first = wb.active
    wb.remove(first)
    for sheet, headers in HUMAN_HEADERS.items():
        ws = wb.create_sheet(sheet)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        if sheet == "GENERAL FOLLOW UP":
            for number, name, owner, status, permit_type in HUMAN_ROWS:
                ws.append([number, name, owner, status, permit_type, "Synthetic human-owned fixture cell"])
        else:
            for number, *_ in HUMAN_ROWS:
                ws.append([number, "R01", "Synthetic", "Synthetic human-owned fixture cell"])
    projection = wb.create_sheet(CANONICAL_PROJECTION_SHEET)
    projection.append(PROJECTION_HEADERS)
    for cell in projection[1]:
        cell.font = Font(bold=True)
    wb.properties.title = "PermitOps Synthetic MVP Dataset v1 — Recording-derived workbook"
    wb.properties.subject = "DEMONSTRATION BASELINE — SYNTHETIC DATA — NOT CLIENT APPROVED"
    wb.save(destination)
    return destination


def canonical_workbook_contract(path: str | Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    return {
        "workbook_identity": str(path),
        "required_sheets": CANONICAL_WORKBOOK_SHEETS,
        "sheets": wb.sheetnames,
        "row_identity": "GENERAL FOLLOW UP: Project Number exact match; sheet + row number retained",
        "human_owned_columns": {sheet: headers for sheet, headers in HUMAN_HEADERS.items()},
        "system_owned_projection": {"sheet": CANONICAL_PROJECTION_SHEET, "columns": PROJECTION_HEADERS[1:]},
        "write_policy": "Only PERMITOPS SYSTEM PROJECTION is writable by the system; human sheets are read-only.",
    }


def _lock_paths(path: Path) -> list[Path]:
    return [path.with_suffix(path.suffix + ".lock"), path.parent / ".permitops-workbook.lock"]


class WorkbookLockedError(RuntimeError):
    pass


class WorkbookWriteRecoveryError(RuntimeError):
    pass


class WorkbookWriteError(RuntimeError):
    pass


def _human_sheet_snapshot(workbook) -> dict[str, tuple[tuple[Any, ...], ...]]:
    return {
        sheet: tuple(
            tuple(cell.value for cell in row)
            for row in workbook[sheet].iter_rows()
        )
        for sheet in HUMAN_HEADERS
    }


def _validate_workbook_candidate(
    path: Path,
    project_number: str,
    values: dict[str, Any],
    human_snapshot: dict[str, tuple[tuple[Any, ...], ...]],
) -> None:
    if not path.is_file() or path.stat().st_size == 0:
        raise WorkbookWriteError("TEMP_WORKBOOK_VALIDATION=FAIL")
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        required_sheets = set(CANONICAL_WORKBOOK_SHEETS) | {CANONICAL_PROJECTION_SHEET}
        if not required_sheets.issubset(set(workbook.sheetnames)):
            raise WorkbookWriteError("TEMP_WORKBOOK_VALIDATION=FAIL")
        if _human_sheet_snapshot(workbook) != human_snapshot:
            raise WorkbookWriteError("HUMAN_SHEET_PRESERVATION=FAIL")
        sheet = workbook[CANONICAL_PROJECTION_SHEET]
        headers = [cell.value for cell in sheet[1]]
        if any(header not in headers for header in PROJECTION_HEADERS):
            raise WorkbookWriteError("PROJECTION_HEADERS=FAIL")
        row = next(
            (
                index
                for index in range(2, sheet.max_row + 1)
                if sheet.cell(index, 1).value == project_number
            ),
        )
        if row is None:
            raise WorkbookWriteError("PROJECTION_ROW=FAIL")
        for key, value in values.items():
            if key in PROJECTION_HEADERS and sheet.cell(row, headers.index(key) + 1).value != value:
                raise WorkbookWriteError("PROJECTION_VALUES=FAIL")
        expected_status = values.get("Projection Status", "WRITTEN")
        if sheet.cell(row, headers.index("Projection Status") + 1).value != expected_status:
            raise WorkbookWriteError("PROJECTION_STATUS=FAIL")
    except WorkbookWriteError:
        raise
    except Exception as exc:
        raise WorkbookWriteError("TEMP_WORKBOOK_VALIDATION=FAIL") from exc
    finally:
        if workbook is not None:
            workbook.close()


def _azure_files_fallback_allowed(destination: Path) -> bool:
    try:
        settings = get_settings()
        configured_root = os.getenv("SYNTHETIC_TEST_ROOT")
        if not configured_root:
            return False
        resolved_destination = destination.resolve()
        return (
            settings.app_env.upper() == "AZURE-PREPROD"
            and settings.synthetic_only is True
            and settings.real_data_allowed is False
            and settings.storage_provider.lower() == "mock"
            and settings.synology_mode.upper() == "SYNTHETIC"
            and resolved_destination == canonical_workbook_path().resolve()
            and resolved_destination.is_relative_to(Path(configured_root).resolve())
        )
    except Exception:
        return False


def _write_full_file(path: Path, content: bytes) -> None:
    with path.open("wb") as handle:
        handle.write(content)
        handle.flush()
        os.fsync(handle.fileno())


def _restore_original_workbook(
    path: Path,
    existed: bool,
    original_bytes: bytes,
    original_sha256: str,
    human_snapshot: dict[str, tuple[tuple[Any, ...], ...]],
) -> None:
    try:
        if existed:
            _write_full_file(path, original_bytes)
            if hashlib.sha256(path.read_bytes()).hexdigest() != original_sha256:
                raise WorkbookWriteRecoveryError("WORKBOOK_RESTORE=FAIL")
            restored = load_workbook(path, read_only=True, data_only=True)
            try:
                if _human_sheet_snapshot(restored) != human_snapshot:
                    raise WorkbookWriteRecoveryError("WORKBOOK_RESTORE=FAIL")
            finally:
                restored.close()
        elif path.exists():
            path.unlink()
    except WorkbookWriteRecoveryError:
        raise
    except Exception as exc:
        raise WorkbookWriteRecoveryError("WORKBOOK_RESTORE=FAIL") from exc


def _bounded_azure_files_overwrite(
    temporary_path: Path,
    destination: Path,
    project_number: str,
    values: dict[str, Any],
    human_snapshot: dict[str, tuple[tuple[Any, ...], ...]],
    original_exists: bool,
    original_bytes: bytes,
    original_sha256: str,
) -> None:
    try:
        _write_full_file(destination, temporary_path.read_bytes())
        _validate_workbook_candidate(destination, project_number, values, human_snapshot)
    except Exception as overwrite_error:
        try:
            _restore_original_workbook(
                destination,
                original_exists,
                original_bytes,
                original_sha256,
                human_snapshot,
            )
        except WorkbookWriteRecoveryError:
            raise
        raise WorkbookWriteError("SYNTHETIC_AZURE_FILES_BOUNDED_OVERWRITE=FAIL") from overwrite_error


def write_system_projection(path: str | Path, project_number: str, values: dict[str, Any]) -> dict[str, Any]:
    workbook_path = Path(path)
    if any(lock.exists() for lock in _lock_paths(workbook_path)):
        raise WorkbookLockedError("WORKBOOK_LOCKED_MANUAL_COPY_REQUIRED")
    wb = load_workbook(workbook_path)
    human_snapshot = _human_sheet_snapshot(wb)
    if CANONICAL_PROJECTION_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError("CANONICAL_PROJECTION_SHEET_MISSING")
    ws = wb[CANONICAL_PROJECTION_SHEET]
    headers = [cell.value for cell in ws[1]]
    row = next((index for index in range(2, ws.max_row + 1) if ws.cell(index, 1).value == project_number), None)
    if row is None:
        row = ws.max_row + 1
        ws.cell(row, 1).value = project_number
    allowed = set(PROJECTION_HEADERS[1:])
    for key, value in values.items():
        if key in allowed:
            ws.cell(row, headers.index(key) + 1).value = value
    ws.cell(row, headers.index("Projection Status") + 1).value = values.get("Projection Status", "WRITTEN")
    with NamedTemporaryFile(prefix="permitops-excel-", suffix=".xlsx", dir=workbook_path.parent, delete=False) as temporary:
        temporary_path = Path(temporary.name)
    try:
        wb.save(temporary_path)
        wb.close()
        _validate_workbook_candidate(temporary_path, project_number, values, human_snapshot)
        try:
            os.replace(temporary_path, workbook_path)
            write_mode = "ATOMIC_REPLACE"
            _validate_workbook_candidate(workbook_path, project_number, values, human_snapshot)
        except PermissionError as replace_error:
            if not _azure_files_fallback_allowed(workbook_path):
                raise
            original_exists = workbook_path.exists()
            original_bytes = workbook_path.read_bytes() if original_exists else b""
            original_sha256 = hashlib.sha256(original_bytes).hexdigest()
            _bounded_azure_files_overwrite(
                temporary_path,
                workbook_path,
                project_number,
                values,
                human_snapshot,
                original_exists,
                original_bytes,
                original_sha256,
            )
            write_mode = "SYNTHETIC_AZURE_FILES_BOUNDED_OVERWRITE"
    finally:
        if temporary_path.exists():
            temporary_path.unlink()
    return {"workbook_identity": str(workbook_path), "sheet": CANONICAL_PROJECTION_SHEET, "row_number": row, "row_key": project_number, "status": "WRITTEN", "write_mode": write_mode, "owned_region_only": True}
